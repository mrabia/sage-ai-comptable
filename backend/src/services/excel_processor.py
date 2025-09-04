import os
import logging
from typing import Optional, Dict, Any, List
import pandas as pd
import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Service pour traiter les fichiers Excel"""
    
    def __init__(self):
        self.supported_extensions = ['.xls', '.xlsx']
        self.max_rows_preview = 1000  # Limite pour l'aperçu
    
    def extract_text(self, file_path: str) -> Optional[str]:
        """Extrait le contenu textuel d'un fichier Excel"""
        try:
            # Lire toutes les feuilles
            excel_file = pd.ExcelFile(file_path)
            text_content = []
            
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=self.max_rows_preview)
                    
                    # Convertir en texte
                    text_content.append(f"=== Feuille: {sheet_name} ===\n")
                    
                    # Ajouter les en-têtes
                    headers = ' | '.join(str(col) for col in df.columns)
                    text_content.append(f"En-têtes: {headers}\n")
                    
                    # Ajouter les données
                    for _, row in df.iterrows():
                        row_text = ' | '.join(str(val) if pd.notna(val) else '' for val in row)
                        text_content.append(f"{row_text}\n")
                    
                    text_content.append("\n")
                    
                except Exception as e:
                    logger.warning(f"Erreur lors de la lecture de la feuille {sheet_name}: {e}")
                    continue
            
            return ''.join(text_content)
            
        except Exception as e:
            logger.error(f"Erreur lors de la lecture du fichier Excel: {e}")
            return None
    
    def extract_structured_data(self, file_path: str) -> Dict[str, Any]:
        """Extrait les données structurées du fichier Excel"""
        try:
            # Analyser le fichier Excel
            workbook_info = self._analyze_workbook(file_path)
            
            structured_data = {
                'workbook_info': workbook_info,
                'sheets': {}
            }
            
            # Traiter chaque feuille
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                try:
                    sheet_data = self._process_sheet(file_path, sheet_name)
                    structured_data['sheets'][sheet_name] = sheet_data
                    
                except Exception as e:
                    logger.warning(f"Erreur lors du traitement de la feuille {sheet_name}: {e}")
                    structured_data['sheets'][sheet_name] = {'error': str(e)}
            
            # Détecter le type de données global
            data_type = self._detect_workbook_data_type(structured_data)
            structured_data['detected_data_type'] = data_type
            
            # Extraire les données spécifiques selon le type
            if data_type == 'clients':
                structured_data['clients_data'] = self._extract_clients_from_workbook(structured_data)
            elif data_type == 'products':
                structured_data['products_data'] = self._extract_products_from_workbook(structured_data)
            elif data_type == 'transactions':
                structured_data['transactions_data'] = self._extract_transactions_from_workbook(structured_data)
            
            return structured_data
            
        except Exception as e:
            logger.error(f"Erreur lors de l'extraction de données structurées: {e}")
            return {'error': str(e)}
    
    def _analyze_workbook(self, file_path: str) -> Dict[str, Any]:
        """Analyse les informations générales du classeur"""
        try:
            info = {}
            
            # Utiliser openpyxl pour les métadonnées
            try:
                workbook = load_workbook(file_path, read_only=True)
                info['sheet_names'] = workbook.sheetnames
                info['num_sheets'] = len(workbook.sheetnames)
                
                # Propriétés du document si disponibles
                props = workbook.properties
                if props:
                    info['title'] = props.title
                    info['creator'] = props.creator
                    info['created'] = str(props.created) if props.created else None
                    info['modified'] = str(props.modified) if props.modified else None
                
                workbook.close()
                
            except Exception as e:
                logger.warning(f"Erreur lors de l'analyse avec openpyxl: {e}")
            
            # Utiliser pandas pour l'analyse des données
            try:
                excel_file = pd.ExcelFile(file_path)
                info['sheet_names'] = excel_file.sheet_names
                info['num_sheets'] = len(excel_file.sheet_names)
                
                # Analyser rapidement chaque feuille
                sheet_summary = {}
                for sheet_name in excel_file.sheet_names:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                        sheet_summary[sheet_name] = {
                            'num_columns': len(df.columns),
                            'column_names': df.columns.tolist(),
                            'has_data': len(df) > 0
                        }
                    except:
                        sheet_summary[sheet_name] = {'error': 'Impossible de lire la feuille'}
                
                info['sheets_summary'] = sheet_summary
                
            except Exception as e:
                logger.warning(f"Erreur lors de l'analyse avec pandas: {e}")
            
            return info
            
        except Exception as e:
            logger.error(f"Erreur lors de l'analyse du classeur: {e}")
            return {}
    
    def _process_sheet(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """Traite une feuille spécifique"""
        try:
            # Lire la feuille
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=self.max_rows_preview)
            
            # Nettoyer les données
            df = self._clean_dataframe(df)
            
            # Analyser la structure
            analysis = self._analyze_sheet_structure(df)
            
            # Détecter le type de données
            data_type = self._detect_sheet_data_type(df)
            
            sheet_data = {
                'data_type': data_type,
                'num_rows': len(df),
                'num_columns': len(df.columns),
                'columns': df.columns.tolist(),
                'column_types': df.dtypes.astype(str).to_dict(),
                'analysis': analysis,
                'preview_data': df.head(10).to_dict('records'),
                'sample_values': self._get_sample_values(df)
            }
            
            return sheet_data
            
        except Exception as e:
            logger.error(f"Erreur lors du traitement de la feuille {sheet_name}: {e}")
            return {'error': str(e)}
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Nettoie le DataFrame"""
        try:
            # Supprimer les lignes complètement vides
            df = df.dropna(how='all')
            
            # Supprimer les colonnes complètement vides
            df = df.dropna(axis=1, how='all')
            
            # Nettoyer les noms de colonnes
            df.columns = df.columns.astype(str).str.strip()
            
            # Remplacer les valeurs vides par None
            df = df.where(pd.notnull(df), None)
            
            return df
            
        except Exception as e:
            logger.warning(f"Erreur lors du nettoyage: {e}")
            return df
    
    def _analyze_sheet_structure(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Analyse la structure de la feuille"""
        try:
            analysis = {}
            
            # Statistiques de base
            analysis['total_cells'] = df.size
            analysis['empty_cells'] = df.isnull().sum().sum()
            analysis['fill_rate'] = 1 - (analysis['empty_cells'] / analysis['total_cells']) if analysis['total_cells'] > 0 else 0
            
            # Analyse des colonnes
            column_analysis = {}
            for column in df.columns:
                col_data = df[column]
                column_analysis[column] = {
                    'non_null_count': col_data.count(),
                    'null_count': col_data.isnull().sum(),
                    'unique_count': col_data.nunique(),
                    'data_type': str(col_data.dtype),
                    'sample_values': col_data.dropna().head(3).tolist()
                }
            
            analysis['columns'] = column_analysis
            
            # Détecter les patterns de données
            analysis['patterns'] = self._detect_data_patterns(df)
            
            return analysis
            
        except Exception as e:
            logger.warning(f"Erreur lors de l'analyse: {e}")
            return {}
    
    def _detect_data_patterns(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Détecte des patterns dans les données"""
        try:
            patterns = {}
            
            # Détecter les colonnes de dates
            date_columns = []
            for column in df.columns:
                if df[column].dtype == 'datetime64[ns]' or 'date' in str(column).lower():
                    date_columns.append(column)
            patterns['date_columns'] = date_columns
            
            # Détecter les colonnes numériques
            numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
            patterns['numeric_columns'] = numeric_columns
            
            # Détecter les colonnes de texte
            text_columns = df.select_dtypes(include=['object']).columns.tolist()
            patterns['text_columns'] = text_columns
            
            # Détecter les colonnes avec des emails
            email_columns = []
            for column in text_columns:
                sample_values = df[column].dropna().head(10).astype(str)
                email_count = sum(1 for val in sample_values if '@' in val and '.' in val)
                if email_count > 0:
                    email_columns.append(column)
            patterns['email_columns'] = email_columns
            
            # Détecter les colonnes avec des montants
            amount_columns = []
            for column in df.columns:
                col_name = str(column).lower()
                if any(keyword in col_name for keyword in ['prix', 'price', 'montant', 'amount', 'total', 'cost']):
                    amount_columns.append(column)
            patterns['amount_columns'] = amount_columns
            
            return patterns
            
        except Exception as e:
            logger.warning(f"Erreur lors de la détection de patterns: {e}")
            return {}
    
    def _detect_sheet_data_type(self, df: pd.DataFrame) -> str:
        """Détecte le type de données de la feuille"""
        try:
            columns = [str(col).lower() for col in df.columns]
            
            # Patterns pour détecter les clients
            client_patterns = [
                'nom', 'name', 'client', 'customer', 'société', 'company',
                'email', 'mail', 'téléphone', 'phone', 'adresse', 'address'
            ]
            
            # Patterns pour détecter les produits
            product_patterns = [
                'produit', 'product', 'article', 'item', 'référence', 'ref',
                'prix', 'price', 'tarif', 'cost', 'tva', 'vat'
            ]
            
            # Patterns pour détecter les transactions
            transaction_patterns = [
                'date', 'montant', 'amount', 'débit', 'crédit', 'debit', 'credit',
                'facture', 'invoice', 'transaction', 'opération'
            ]
            
            # Compter les correspondances
            client_score = sum(1 for pattern in client_patterns 
                             if any(pattern in col for col in columns))
            
            product_score = sum(1 for pattern in product_patterns 
                              if any(pattern in col for col in columns))
            
            transaction_score = sum(1 for pattern in transaction_patterns 
                                  if any(pattern in col for col in columns))
            
            # Déterminer le type
            if client_score >= 2:
                return 'clients'
            elif product_score >= 2:
                return 'products'
            elif transaction_score >= 2:
                return 'transactions'
            else:
                return 'generic'
                
        except Exception as e:
            logger.warning(f"Erreur lors de la détection du type: {e}")
            return 'generic'
    
    def _detect_workbook_data_type(self, structured_data: Dict[str, Any]) -> str:
        """Détecte le type de données global du classeur"""
        try:
            sheet_types = []
            
            for sheet_name, sheet_data in structured_data.get('sheets', {}).items():
                if 'data_type' in sheet_data:
                    sheet_types.append(sheet_data['data_type'])
            
            if not sheet_types:
                return 'generic'
            
            # Prendre le type le plus fréquent
            type_counts = {}
            for data_type in sheet_types:
                type_counts[data_type] = type_counts.get(data_type, 0) + 1
            
            return max(type_counts, key=type_counts.get)
            
        except Exception as e:
            logger.warning(f"Erreur lors de la détection du type global: {e}")
            return 'generic'
    
    def _get_sample_values(self, df: pd.DataFrame) -> Dict[str, List]:
        """Récupère des valeurs d'exemple pour chaque colonne"""
        try:
            sample_values = {}
            
            for column in df.columns:
                # Prendre 5 valeurs non-nulles uniques
                values = df[column].dropna().unique()[:5].tolist()
                sample_values[column] = values
            
            return sample_values
            
        except Exception as e:
            logger.warning(f"Erreur lors de l'extraction des valeurs d'exemple: {e}")
            return {}
    
    def _extract_clients_from_workbook(self, structured_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extrait les données clients du classeur"""
        try:
            all_clients = []
            
            for sheet_name, sheet_data in structured_data.get('sheets', {}).items():
                if sheet_data.get('data_type') == 'clients' and 'preview_data' in sheet_data:
                    # Mapper les colonnes pour cette feuille
                    columns = sheet_data.get('columns', [])
                    column_mapping = self._map_client_columns(columns)
                    
                    # Extraire les clients de cette feuille
                    for row in sheet_data['preview_data']:
                        client = {}
                        
                        for standard_field, csv_column in column_mapping.items():
                            if csv_column and csv_column in row:
                                value = row[csv_column]
                                if value is not None:
                                    client[standard_field] = str(value).strip()
                        
                        if client:
                            client['source_sheet'] = sheet_name
                            all_clients.append(client)
            
            return {
                'clients': all_clients,
                'total_count': len(all_clients),
                'source_sheets': list(set(client.get('source_sheet') for client in all_clients))
            }
            
        except Exception as e:
            logger.error(f"Erreur lors de l'extraction des clients: {e}")
            return {'error': str(e)}
    
    def _extract_products_from_workbook(self, structured_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extrait les données produits du classeur"""
        try:
            all_products = []
            
            for sheet_name, sheet_data in structured_data.get('sheets', {}).items():
                if sheet_data.get('data_type') == 'products' and 'preview_data' in sheet_data:
                    # Mapper les colonnes pour cette feuille
                    columns = sheet_data.get('columns', [])
                    column_mapping = self._map_product_columns(columns)
                    
                    # Extraire les produits de cette feuille
                    for row in sheet_data['preview_data']:
                        product = {}
                        
                        for standard_field, csv_column in column_mapping.items():
                            if csv_column and csv_column in row:
                                value = row[csv_column]
                                if value is not None:
                                    product[standard_field] = str(value).strip()
                        
                        if product:
                            product['source_sheet'] = sheet_name
                            all_products.append(product)
            
            return {
                'products': all_products,
                'total_count': len(all_products),
                'source_sheets': list(set(product.get('source_sheet') for product in all_products))
            }
            
        except Exception as e:
            logger.error(f"Erreur lors de l'extraction des produits: {e}")
            return {'error': str(e)}
    
    def _extract_transactions_from_workbook(self, structured_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extrait les données de transactions du classeur"""
        try:
            all_transactions = []
            
            for sheet_name, sheet_data in structured_data.get('sheets', {}).items():
                if sheet_data.get('data_type') == 'transactions' and 'preview_data' in sheet_data:
                    # Mapper les colonnes pour cette feuille
                    columns = sheet_data.get('columns', [])
                    column_mapping = self._map_transaction_columns(columns)
                    
                    # Extraire les transactions de cette feuille
                    for row in sheet_data['preview_data']:
                        transaction = {}
                        
                        for standard_field, csv_column in column_mapping.items():
                            if csv_column and csv_column in row:
                                value = row[csv_column]
                                if value is not None:
                                    transaction[standard_field] = str(value).strip()
                        
                        if transaction:
                            transaction['source_sheet'] = sheet_name
                            all_transactions.append(transaction)
            
            return {
                'transactions': all_transactions,
                'total_count': len(all_transactions),
                'source_sheets': list(set(transaction.get('source_sheet') for transaction in all_transactions))
            }
            
        except Exception as e:
            logger.error(f"Erreur lors de l'extraction des transactions: {e}")
            return {'error': str(e)}
    
    def _map_client_columns(self, columns: List[str]) -> Dict[str, Optional[str]]:
        """Mappe les colonnes vers les champs clients standards"""
        # Réutiliser la logique du CSV processor
        from src.services.csv_processor import CSVProcessor
        csv_processor = CSVProcessor()
        return csv_processor._map_client_columns(columns)
    
    def _map_product_columns(self, columns: List[str]) -> Dict[str, Optional[str]]:
        """Mappe les colonnes vers les champs produits standards"""
        # Réutiliser la logique du CSV processor
        from src.services.csv_processor import CSVProcessor
        csv_processor = CSVProcessor()
        return csv_processor._map_product_columns(columns)
    
    def _map_transaction_columns(self, columns: List[str]) -> Dict[str, Optional[str]]:
        """Mappe les colonnes vers les champs transactions standards"""
        # Réutiliser la logique du CSV processor
        from src.services.csv_processor import CSVProcessor
        csv_processor = CSVProcessor()
        return csv_processor._map_transaction_columns(columns)
    
    def is_valid_excel(self, file_path: str) -> bool:
        """Vérifie si le fichier est un Excel valide"""
        try:
            # Tenter de lire le fichier
            excel_file = pd.ExcelFile(file_path)
            
            # Vérifier qu'il y a au moins une feuille
            if len(excel_file.sheet_names) == 0:
                return False
            
            # Tenter de lire la première feuille
            first_sheet = excel_file.sheet_names[0]
            df = pd.read_excel(file_path, sheet_name=first_sheet, nrows=5)
            
            # Vérifier qu'on a des données
            return len(df.columns) > 0
            
        except Exception as e:
            logger.error(f"Excel invalide: {e}")
            return False

